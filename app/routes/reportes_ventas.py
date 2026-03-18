# app/routes/reportes_ventas.py
# Reportes de resumen: panel de KPIs y exportaciones por fecha/vendedor

from datetime import date, datetime
from io import BytesIO
from flask import Blueprint, render_template, request, send_file, flash, redirect, url_for
from flask_login import login_required, current_user
from sqlalchemy import case, literal
from openpyxl.utils.cell import get_column_letter
import pandas as pd

from app import db
from app.utils.roles import rol_requerido
from app.utils.fechas import contar_habiles, dias_habiles_mes, parsear_fecha
from app.models.vendedor import Vendedor
from app.models.ventas import BDVenta
from app.models.pedidos import BDPedido
from app.models.pedido_item import BDPedidoItem

reportes_ventas_bp = Blueprint('reportes_ventas', __name__, url_prefix='/reportes')


@reportes_ventas_bp.route('/panel', methods=['GET'])
@login_required
@rol_requerido('semiadmin', 'administrador')
def panel():
    hoy = date.today()
    inicio_mes = hoy.replace(day=1)

    dias_transcurridos = contar_habiles(inicio_mes, hoy)
    total_habiles      = dias_habiles_mes(hoy.year, hoy.month)

    rows = []
    for vend in Vendedor.query.order_by(Vendedor.nombre).all():
        ventas = BDVenta.query.filter(
            BDVenta.codigo_vendedor == vend.codigo_vendedor,
            BDVenta.fecha >= inicio_mes,
            BDVenta.fecha <= hoy
        ).all()

        tot_vendido = sum(v.total_venta for v in ventas)
        tot_com     = sum(v.comision    for v in ventas)
        proy = (tot_vendido / dias_transcurridos * total_habiles
                if dias_transcurridos else 0)

        rows.append({
            'vendedor':   vend.nombre,
            'cod_vend':   vend.codigo_vendedor,
            'vendido':    tot_vendido,
            'comision':   tot_com,
            'proyeccion': proy
        })

    return render_template(
        'reportes/panel.html',
        rows=rows,
        dias_transcurridos=dias_transcurridos,
        total_habiles=total_habiles
    )


@reportes_ventas_bp.route('/mi_panel', methods=['GET'])
@login_required
@rol_requerido('vendedor')
def mi_panel():
    from app.models.extras import BDExtra
    from app.models.cambio import BD_CAMBIO

    hoy = date.today()

    mes  = int(request.args.get('mes',  hoy.month))
    anio = int(request.args.get('anio', hoy.year))

    inicio_mes = date(anio, mes, 1)
    fin_mes = date(anio + 1, 1, 1) if mes == 12 else date(anio, mes + 1, 1)

    dias_transcurridos = contar_habiles(inicio_mes, hoy if (anio == hoy.year and mes == hoy.month) else fin_mes)
    total_habiles      = dias_habiles_mes(anio, mes)

    ventas = BDVenta.query.filter(
        BDVenta.codigo_vendedor == current_user.codigo_vendedor,
        BDVenta.fecha >= inicio_mes,
        BDVenta.fecha < fin_mes
    ).all()

    tot_vendido = sum(v.total_venta for v in ventas)
    tot_com     = sum(v.comision    for v in ventas)
    proy = (tot_vendido / dias_transcurridos * total_habiles) if dias_transcurridos else 0

    pedidos = BDPedido.query.filter_by(codigo_vendedor=current_user.codigo_vendedor).filter(
        BDPedido.fecha >= inicio_mes,
        BDPedido.fecha < fin_mes
    ).all()
    total_pedidos = sum(sum(item.subtotal for item in p.items) for p in pedidos)

    extras = BDExtra.query.filter_by(codigo_vendedor=current_user.codigo_vendedor).filter(
        BDExtra.fecha >= inicio_mes,
        BDExtra.fecha < fin_mes
    ).all()
    total_extras = sum(sum(item.subtotal for item in e.items) for e in extras)
    total_pedidos += total_extras

    cambios = BD_CAMBIO.query.filter_by(codigo_vendedor=current_user.codigo_vendedor).filter(
        BD_CAMBIO.fecha >= inicio_mes,
        BD_CAMBIO.fecha < fin_mes
    ).all()
    total_cambios = sum(c.valor_cambio for c in cambios)

    return render_template(
        'reportes/mi_panel.html',
        anio=anio,
        mes=mes,
        vendido=tot_vendido,
        comision=tot_com,
        proyeccion=proy,
        total_pedidos=total_pedidos,
        total_cambios=total_cambios,
        dias_transcurridos=dias_transcurridos,
        total_habiles=total_habiles
    )


# 3) Formulario para pedidos día a día por vendedor
@reportes_ventas_bp.route('/pedidos_dia', methods=['GET'])
@login_required
@rol_requerido('administrador', 'semiadmin')
def pedidos_dia_por_vendedor():
    return render_template('reportes/pedidos_dia_por_vendedor.html')


# 4) Exportación a Excel de pedidos día a día
@reportes_ventas_bp.route('/pedidos_dia/export', methods=['GET'])
@login_required
@rol_requerido('administrador', 'semiadmin')
def export_pedidos_dia_excel():
    start = request.args.get('start', '').strip()
    end   = request.args.get('end',   '').strip()
    try:
        dt_start = parsear_fecha(start)
        dt_end   = parsear_fecha(end)
    except ValueError as e:
        flash(str(e), 'danger')
        return redirect(url_for('reportes_ventas.pedidos_dia_por_vendedor'))

    q = (
        db.session.query(
            BDPedido.fecha.label('Fecha'),
            Vendedor.nombre.label('Vendedor'),
            db.func.sum(BDPedidoItem.subtotal).label('Total')
        )
        .join(BDPedidoItem, BDPedidoItem.pedido_id == BDPedido.id)
        .join(Vendedor, Vendedor.codigo_vendedor == BDPedido.codigo_vendedor)
        .filter(BDPedido.fecha >= dt_start, BDPedido.fecha <= dt_end)
        .group_by(BDPedido.fecha, Vendedor.nombre)
        .order_by(BDPedido.fecha, Vendedor.nombre)
    )

    rows = [
        {'Fecha': rec.Fecha, 'Vendedor': rec.Vendedor, 'Total': float(rec.Total)}
        for rec in q
    ]
    if not rows:
        flash('No hay datos para ese rango.', 'info')
        return redirect(url_for('reportes_ventas.pedidos_dia_por_vendedor'))

    df = pd.DataFrame(rows)
    df['Fecha'] = pd.to_datetime(df['Fecha'])
    table = df.pivot(index='Fecha', columns='Vendedor', values='Total').fillna(0)

    table.loc['Total'] = table.sum(axis=0)
    dias_hab = contar_habiles(dt_start, dt_end)
    tot_hab  = dias_habiles_mes(dt_start.year, dt_start.month)
    proj = table.loc['Total'] / dias_hab * tot_hab if dias_hab else table.loc['Total'] * 0
    table.loc['Proyección'] = proj

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        table.to_excel(writer, sheet_name='PedidosDia')
        ws = writer.sheets['PedidosDia']
        df_reset = table.reset_index()
        for idx, col in enumerate(df_reset.columns, start=1):
            max_len = max(df_reset[col].astype(str).map(len).max(), len(str(col)))
            ws.column_dimensions[get_column_letter(idx)].width = max_len + 2
    output.seek(0)

    filename = f"PedidosDia_{start}_a_{end}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


# 5) Formulario para pedidos mes a mes por vendedor
@reportes_ventas_bp.route('/pedidos_mes', methods=['GET'])
@login_required
@rol_requerido('administrador', 'semiadmin')
def pedidos_mes_por_vendedor():
    return render_template('reportes/pedidos_mes_por_vendedor.html', now=date.today)


# 6) Exportación a Excel de pedidos mes a mes
@reportes_ventas_bp.route('/pedidos_mes/export', methods=['GET'])
@login_required
@rol_requerido('administrador', 'semiadmin')
def export_pedidos_mes_excel():
    year_str = request.args.get('year', '').strip()
    try:
        year = int(year_str)
    except ValueError:
        flash('Año inválido.', 'danger')
        return redirect(url_for('reportes_ventas.pedidos_mes_por_vendedor'))

    q = (
        db.session.query(
            db.func.extract('month', BDPedido.fecha).label('Mes'),
            Vendedor.nombre.label('Vendedor'),
            db.func.sum(BDPedidoItem.subtotal).label('Total')
        )
        .join(BDPedidoItem, BDPedidoItem.pedido_id == BDPedido.id)
        .join(Vendedor, Vendedor.codigo_vendedor == BDPedido.codigo_vendedor)
        .filter(db.func.extract('year', BDPedido.fecha) == year)
        .group_by('Mes', 'Vendedor')
        .order_by('Mes', 'Vendedor')
    )

    rows = [
        {'Mes': int(mes), 'Vendedor': vendedor, 'Total': float(total)}
        for mes, vendedor, total in q
    ]
    if not rows:
        flash(f'No hay datos para el año {year}.', 'info')
        return redirect(url_for('reportes_ventas.pedidos_mes_por_vendedor'))

    df = pd.DataFrame(rows)
    table = df.pivot(index='Mes', columns='Vendedor', values='Total').fillna(0)

    table.loc['Total'] = table.sum(axis=0)
    meses_pasados = date.today().month if date.today().year == year else 12
    proj_values = table.loc['Total'] / meses_pasados * 12 if meses_pasados else table.loc['Total'] * 0
    table.loc['Proyección'] = proj_values

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        table.to_excel(writer, sheet_name=f'Pedidos_{year}')
        ws = writer.sheets[f'Pedidos_{year}']
        df_reset = table.reset_index()
        for idx, col in enumerate(df_reset.columns, start=1):
            max_len = max(df_reset[col].astype(str).map(len).max(), len(str(col)))
            ws.column_dimensions[get_column_letter(idx)].width = max_len + 2
    output.seek(0)

    filename = f"PedidosMes_{year}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
