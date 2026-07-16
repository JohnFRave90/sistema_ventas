# app/routes/reportes_comisiones.py
# Reportes de detalle por producto con cálculo de comisiones

from datetime import date
from io import BytesIO
from decimal import Decimal
from flask import Blueprint, render_template, request, send_file, flash, redirect, url_for
from flask_login import login_required
from openpyxl.utils.cell import get_column_letter
import pandas as pd

from app import db
from app.utils.roles import rol_requerido
from app.utils.fechas import parsear_fecha
from app.utils.queries import obtener_mapa_vendedores, obtener_mapa_vendedores_obj
from app.models.ventas import BDVenta
from app.models.venta_item import BDVentaItem
from app.models.pedidos import BDPedido
from app.models.pedido_item import BDPedidoItem
from app.models.extras import BDExtra
from app.models.extra_item import BDExtraItem
from app.models.devoluciones import BDDevolucion
from app.models.devolucion_item import BDDevolucionItem
from app.models.vendedor import Vendedor
from app.models.producto import Producto

reportes_comisiones_bp = Blueprint('reportes_comisiones', __name__, url_prefix='/reportes')


# 1) Formulario para pedidos por producto
@reportes_comisiones_bp.route('/pedidos_por_producto', methods=['GET'])
@login_required
@rol_requerido('administrador', 'semiadmin')
def pedidos_por_producto():
    return render_template('reportes/pedidos_por_producto.html')


# 2) Exportación a Excel de pedidos por producto
@reportes_comisiones_bp.route('/pedidos_por_producto/export', methods=['GET'])
@login_required
@rol_requerido('administrador', 'semiadmin')
def export_pedidos_productos_excel():
    start = request.args.get('start', '').strip()
    end   = request.args.get('end',   '').strip()
    try:
        start_date = parsear_fecha(start)
        end_date   = parsear_fecha(end)
    except ValueError as e:
        flash(str(e), 'warning')
        return redirect(url_for('reportes_comisiones.pedidos_por_producto'))

    rows = []
    pedidos = BDPedido.query.filter(
        BDPedido.fecha >= start_date,
        BDPedido.fecha <= end_date
    ).all()
    vend_map = obtener_mapa_vendedores_obj()

    for ped in pedidos:
        vend = vend_map.get(ped.codigo_vendedor)
        for item in ped.items:
            prod = Producto.query.filter_by(codigo=item.producto_cod).first()
            cat = (prod.categoria or '').lower()
            pct = ((vend.comision_panaderia if cat == 'panadería' else vend.comision_bizcocheria) or 0) / 100.0
            valor_total = float(item.subtotal)
            valor_neto  = valor_total * (1.0 - pct)

            rows.append({
                'Fecha':                       ped.fecha,
                'Año':                         ped.fecha.year,
                'cod vendedor':                ped.codigo_vendedor,
                'nombre vendedor':             vend.nombre,
                'ruta':                        '',
                'codigo producto':             item.producto_cod,
                'nombre producto':             prod.nombre,
                'cantidad':                    item.cantidad,
                'valor total':                 valor_total,
                'valor neto (menos comisión)': valor_neto,
                'lote':                        '',
                'mes':                         ped.fecha.month,
                'día':                         ped.fecha.day,
                'nombre del día':              ped.fecha.strftime('%A')
            })

    df = pd.DataFrame(rows)
    if df.empty:
        flash('No hay datos en ese rango.', 'info')
        return redirect(url_for('reportes_comisiones.pedidos_por_producto'))
    df['Fecha'] = pd.to_datetime(df['Fecha'])

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='PedidosPorProducto')
        ws = writer.sheets['PedidosPorProducto']
        for idx, col in enumerate(df.columns, start=1):
            max_len = max(df[col].astype(str).map(len).max(), len(col))
            ws.column_dimensions[get_column_letter(idx)].width = max_len + 2
    output.seek(0)

    filename = f"PedidosPorProducto_{start}_a_{end}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


# Formulario de extras por producto
@reportes_comisiones_bp.route('/extra_por_producto', methods=['GET'])
@login_required
@rol_requerido('administrador', 'semiadmin')
def extras_por_producto():
    return render_template('reportes/extra_por_producto.html')


# Exportación a Excel de extras por producto
@reportes_comisiones_bp.route('/extra_por_producto/export', methods=['GET'])
@login_required
@rol_requerido('administrador', 'semiadmin')
def export_extras_productos_excel():
    start = request.args.get('start', '').strip()
    end   = request.args.get('end',   '').strip()
    try:
        start_date = parsear_fecha(start)
        end_date   = parsear_fecha(end)
    except ValueError as e:
        flash(str(e), 'warning')
        return redirect(url_for('reportes_comisiones.extras_por_producto'))

    rows = []
    extras = BDExtra.query.filter(
        BDExtra.fecha >= start_date,
        BDExtra.fecha <= end_date
    ).all()
    vend_map = obtener_mapa_vendedores_obj()

    for ex in extras:
        vend = vend_map.get(ex.codigo_vendedor)
        for item in ex.items:
            prod = Producto.query.filter_by(codigo=item.producto_cod).first()
            cat = (prod.categoria or '').lower()
            pct = ((vend.comision_panaderia if cat == 'panadería' else vend.comision_bizcocheria) or 0) / 100.0
            valor_total = float(item.subtotal)
            valor_neto  = valor_total * (1.0 - pct)

            rows.append({
                'Fecha':                       ex.fecha,
                'Año':                         ex.fecha.year,
                'cod vendedor':                ex.codigo_vendedor,
                'nombre vendedor':             vend.nombre,
                'ruta':                        '',
                'codigo producto':             item.producto_cod,
                'nombre producto':             prod.nombre,
                'cantidad':                    item.cantidad,
                'valor total':                 valor_total,
                'valor neto (menos comisión)': valor_neto,
                'lote':                        '',
                'mes':                         ex.fecha.month,
                'día':                         ex.fecha.day,
                'nombre del día':              ex.fecha.strftime('%A')
            })

    df = pd.DataFrame(rows)
    if df.empty:
        flash('No hay datos en ese rango.', 'info')
        return redirect(url_for('reportes_comisiones.extras_por_producto'))
    df['Fecha'] = pd.to_datetime(df['Fecha'])

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='ExtrasPorProducto')
        ws = writer.sheets['ExtrasPorProducto']
        for idx, col in enumerate(df.columns, start=1):
            max_len = max(df[col].astype(str).map(len).max(), len(col))
            ws.column_dimensions[get_column_letter(idx)].width = max_len + 2
    output.seek(0)

    filename = f"ExtrasPorProducto_{start}_a_{end}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# Formulario de devoluciones por producto
@reportes_comisiones_bp.route('/devoluciones_por_producto', methods=['GET'])
@login_required
@rol_requerido('administrador', 'semiadmin')
def devoluciones_por_producto():
    return render_template('reportes/devoluciones_por_producto.html')


# Exportación a Excel de devoluciones por producto
@reportes_comisiones_bp.route('/devoluciones_por_producto/export', methods=['GET'])
@login_required
@rol_requerido('administrador', 'semiadmin')
def export_devoluciones_productos_excel():
    start = request.args.get('start', '').strip()
    end   = request.args.get('end',   '').strip()
    try:
        start_date = parsear_fecha(start)
        end_date   = parsear_fecha(end)
    except ValueError as e:
        flash(str(e), 'warning')
        return redirect(url_for('reportes_comisiones.devoluciones_por_producto'))

    rows = []
    devols = BDDevolucion.query.filter(
        BDDevolucion.fecha >= start_date,
        BDDevolucion.fecha <= end_date
    ).all()
    vend_map = obtener_mapa_vendedores_obj()

    for dev in devols:
        vend = vend_map.get(dev.codigo_vendedor)
        for item in dev.items:
            prod = Producto.query.filter_by(codigo=item.producto_cod).first()
            cat = (prod.categoria or '').lower()
            pct = ((vend.comision_panaderia if cat == 'panadería' else vend.comision_bizcocheria) or 0) / 100.0
            valor_total = float(item.subtotal)
            valor_neto  = valor_total * (1.0 - pct)

            rows.append({
                'Fecha':                       dev.fecha,
                'Año':                         dev.fecha.year,
                'cod vendedor':                dev.codigo_vendedor,
                'nombre vendedor':             vend.nombre,
                'ruta':                        '',
                'codigo producto':             item.producto_cod,
                'nombre producto':             prod.nombre,
                'cantidad':                    item.cantidad,
                'valor total':                 valor_total,
                'valor neto (menos comisión)': valor_neto,
                'lote':                        '',
                'mes':                         dev.fecha.month,
                'día':                         dev.fecha.day,
                'nombre del día':              dev.fecha.strftime('%A')
            })

    df = pd.DataFrame(rows)
    if df.empty:
        flash('No hay datos en ese rango.', 'info')
        return redirect(url_for('reportes_comisiones.devoluciones_por_producto'))
    df['Fecha'] = pd.to_datetime(df['Fecha'])

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='DevolucionesPorProducto')
        ws = writer.sheets['DevolucionesPorProducto']
        for idx, col in enumerate(df.columns, start=1):
            max_len = max(df[col].astype(str).map(len).max(), len(col))
            ws.column_dimensions[get_column_letter(idx)].width = max_len + 2
    output.seek(0)

    filename = f"DevolucionesPorProducto_{start}_a_{end}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# Formulario para ventas por producto
@reportes_comisiones_bp.route('/ventas_por_producto', methods=['GET'])
@login_required
@rol_requerido('administrador', 'semiadmin')
def ventas_por_producto():
    return render_template('reportes/ventas_por_producto.html')


# Exportación a Excel de ventas por producto
@reportes_comisiones_bp.route('/ventas_por_producto/export', methods=['GET'])
@login_required
@rol_requerido('administrador', 'semiadmin')
def export_ventas_producto_excel():
    start = request.args.get('start', '').strip()
    end   = request.args.get('end',   '').strip()

    try:
        start_date = parsear_fecha(start)
        end_date   = parsear_fecha(end)
    except ValueError as e:
        flash(str(e), 'warning')
        return redirect(url_for('reportes_comisiones.ventas_por_producto'))

    ventas = BDVenta.query.filter(
        BDVenta.fecha >= start_date,
        BDVenta.fecha <= end_date
    ).all()
    nombre_vend_map = obtener_mapa_vendedores()

    rows = []
    for v in ventas:
        vendedor_nombre = nombre_vend_map.get(v.codigo_vendedor, 'Desconocido')
        for item in v.items:
            producto = Producto.query.filter_by(codigo=item.producto_cod).first()
            comision_valor  = item.comision  or Decimal('0')
            pagar_panaderia = item.pagar_pan or Decimal('0')

            rows.append({
                'Fecha':                 v.fecha,
                'Año':                   v.fecha.year,
                'Mes':                   v.fecha.month,
                'Día':                   v.fecha.day,
                'Día de semana':         v.fecha.strftime('%A'),
                'Código producto':       item.producto_cod,
                'Nombre producto':       producto.nombre if producto else 'Desconocido',
                'Cantidad':              item.cantidad,
                'Subtotal':              float(item.subtotal),
                'Valor comisión':        float(item.comision),
                'Pagar a la Panadería':  float(item.pagar_pan),
                'Código vendedor':       v.codigo_vendedor,
                'Nombre vendedor':       vendedor_nombre
            })

    df = pd.DataFrame(rows)
    if df.empty:
        flash('No hay datos en ese rango.', 'info')
        return redirect(url_for('reportes_comisiones.ventas_por_producto'))

    df['Fecha'] = pd.to_datetime(df['Fecha'])

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='VentasPorProducto')
        ws = writer.sheets['VentasPorProducto']
        for idx, col in enumerate(df.columns, start=1):
            max_len = max(df[col].astype(str).map(len).max(), len(col))
            ws.column_dimensions[get_column_letter(idx)].width = max_len + 2
    output.seek(0)

    filename = f"VentasPorProducto_{start}_a_{end}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
