from flask import Blueprint, render_template, request, redirect, url_for, flash, Response, send_file, jsonify
from datetime import date
import csv
import io
from app.models.cliente import Cliente
from app.models.vendedor import Vendedor
from app import db
from flask_login import login_required
from app.utils.roles import rol_requerido
from app.services.mapa_service import colores_por_vendedor, COLOR_VENDEDOR

clientes_bp = Blueprint('clientes', __name__, url_prefix='/clientes')

LATITUD_PANADERIA = 6.98993839765757
LONGITUD_PANADERIA = -73.0472798998108


@clientes_bp.route('/')
@login_required
@rol_requerido('administrador')
def listar_clientes():
    vendedor_filtro = request.args.get('vendedor', '')
    query = Cliente.query
    if vendedor_filtro:
        query = query.filter_by(codigo_vendedor=vendedor_filtro)
    clientes = query.order_by(Cliente.codigo_vendedor, Cliente.ruta, Cliente.orden_visita).all()
    vendedores = Vendedor.query.order_by(Vendedor.nombre).all()

    def _ubicacion_pendiente(c):
        if c.latitud is None or c.longitud is None:
            return False
        return (
            abs(float(c.latitud) - LATITUD_PANADERIA) < 0.0001
            and abs(float(c.longitud) - LONGITUD_PANADERIA) < 0.0001
        )

    ids_ubicacion_pendiente = {c.id for c in clientes if _ubicacion_pendiente(c)}

    color_por_vendedor = colores_por_vendedor()
    clientes_geo = [
        {
            'nombre': c.nombre,
            'nombre_tienda': c.nombre_tienda or '',
            'telefono': c.telefono or '',
            'direccion': c.direccion or '',
            'barrio': c.barrio or '',
            'vendedor': c.codigo_vendedor,
            'color': color_por_vendedor.get(c.codigo_vendedor, COLOR_VENDEDOR),
            'lat': float(c.latitud),
            'lng': float(c.longitud),
            'pendiente': c.id in ids_ubicacion_pendiente,
        }
        for c in clientes
        if c.latitud is not None and c.longitud is not None
    ]
    leyenda_vendedores = [
        {'codigo': v.codigo_vendedor, 'nombre': v.nombre, 'color': color_por_vendedor.get(v.codigo_vendedor, COLOR_VENDEDOR)}
        for v in vendedores
        if any(cg['vendedor'] == v.codigo_vendedor for cg in clientes_geo)
    ]

    return render_template(
        'clientes/listar.html',
        clientes=clientes,
        vendedores=vendedores,
        vendedor_filtro=vendedor_filtro,
        clientes_geo=clientes_geo,
        leyenda_vendedores=leyenda_vendedores,
        ids_ubicacion_pendiente=ids_ubicacion_pendiente,
    )


@clientes_bp.route('/crear', methods=['GET', 'POST'])
@login_required
@rol_requerido('administrador')
def crear_cliente():
    vendedores = Vendedor.query.order_by(Vendedor.nombre).all()

    if request.method == 'POST':
        nombre = (request.form.get('nombre') or '').strip()
        codigo_vendedor = (request.form.get('codigo_vendedor') or '').strip()

        if not nombre or not codigo_vendedor:
            flash('Nombre y vendedor son requeridos.', 'danger')
            return render_template('clientes/crear.html', vendedores=vendedores)

        nuevo = Cliente(
            codigo_cliente=Cliente.siguiente_codigo(),
            nombre=nombre,
            nombre_tienda=(request.form.get('nombre_tienda') or '').strip() or None,
            codigo_vendedor=codigo_vendedor,
            telefono=(request.form.get('telefono') or '').strip() or None,
            whatsapp=(request.form.get('whatsapp') or '').strip() or None,
            direccion=(request.form.get('direccion') or '').strip() or None,
            barrio=(request.form.get('barrio') or '').strip() or None,
            ciudad=(request.form.get('ciudad') or '').strip() or None,
            ruta=(request.form.get('ruta') or '').strip() or None,
        )
        lat = (request.form.get('latitud') or '').strip()
        lng = (request.form.get('longitud') or '').strip()
        orden = (request.form.get('orden_visita') or '').strip()
        try:
            nuevo.latitud = float(lat) if lat else LATITUD_PANADERIA
        except ValueError:
            nuevo.latitud = LATITUD_PANADERIA
        try:
            nuevo.longitud = float(lng) if lng else LONGITUD_PANADERIA
        except ValueError:
            nuevo.longitud = LONGITUD_PANADERIA
        if orden:
            try:
                nuevo.orden_visita = int(orden)
            except ValueError:
                pass

        db.session.add(nuevo)
        db.session.commit()
        flash(f'Cliente {nuevo.codigo_cliente} creado.', 'success')
        return redirect(url_for('clientes.listar_clientes'))

    return render_template('clientes/crear.html', vendedores=vendedores)


@clientes_bp.route('/<int:id>/editar', methods=['GET', 'POST'])
@login_required
@rol_requerido('administrador')
def editar_cliente(id):
    cliente = Cliente.query.get_or_404(id)
    vendedores = Vendedor.query.order_by(Vendedor.nombre).all()

    if request.method == 'POST':
        nombre = (request.form.get('nombre') or '').strip()
        codigo_vendedor = (request.form.get('codigo_vendedor') or '').strip()

        if not nombre or not codigo_vendedor:
            flash('Nombre y vendedor son requeridos.', 'danger')
            return render_template('clientes/editar.html', cliente=cliente, vendedores=vendedores)

        cliente.nombre = nombre
        cliente.nombre_tienda = (request.form.get('nombre_tienda') or '').strip() or None
        cliente.codigo_vendedor = codigo_vendedor
        cliente.telefono = (request.form.get('telefono') or '').strip() or None
        cliente.whatsapp = (request.form.get('whatsapp') or '').strip() or None
        cliente.direccion = (request.form.get('direccion') or '').strip() or None
        cliente.barrio = (request.form.get('barrio') or '').strip() or None
        cliente.ciudad = (request.form.get('ciudad') or '').strip() or None
        cliente.ruta = (request.form.get('ruta') or '').strip() or None
        cliente.activo = 'activo' in request.form

        lat = (request.form.get('latitud') or '').strip()
        lng = (request.form.get('longitud') or '').strip()
        orden = (request.form.get('orden_visita') or '').strip()
        cliente.latitud = float(lat) if lat else LATITUD_PANADERIA
        cliente.longitud = float(lng) if lng else LONGITUD_PANADERIA
        cliente.orden_visita = int(orden) if orden else None

        db.session.commit()
        flash('Cliente actualizado.', 'success')
        return redirect(url_for('clientes.listar_clientes'))

    return render_template('clientes/editar.html', cliente=cliente, vendedores=vendedores)


@clientes_bp.route('/<int:id>/toggle', methods=['POST'])
@login_required
@rol_requerido('administrador')
def toggle_cliente(id):
    cliente = Cliente.query.get_or_404(id)
    cliente.activo = not cliente.activo
    db.session.commit()
    estado = 'activado' if cliente.activo else 'desactivado'
    flash(f'Cliente {cliente.nombre} {estado}.', 'success')
    return redirect(url_for('clientes.listar_clientes'))


@clientes_bp.route('/<int:id>/toggle-ajax', methods=['POST'])
@login_required
@rol_requerido('administrador')
def toggle_cliente_ajax(id):
    cliente = Cliente.query.get_or_404(id)
    cliente.activo = not cliente.activo
    db.session.commit()
    return jsonify({'ok': True, 'activo': cliente.activo})


@clientes_bp.route('/activar-lote', methods=['POST'])
@login_required
@rol_requerido('administrador')
def activar_lote_clientes():
    data = request.get_json(silent=True) or {}
    try:
        ids = [int(i) for i in (data.get('ids') or [])]
    except (TypeError, ValueError):
        return jsonify({'ok': False, 'error': 'ids inválidos'}), 400
    activo = bool(data.get('activo'))
    if not ids:
        return jsonify({'ok': False, 'error': 'No se seleccionó ningún cliente'}), 400

    actualizados = Cliente.query.filter(Cliente.id.in_(ids)).update(
        {'activo': activo}, synchronize_session=False
    )
    db.session.commit()
    return jsonify({'ok': True, 'actualizados': actualizados, 'activo': activo})


@clientes_bp.route('/ordenar', methods=['GET'])
@login_required
@rol_requerido('administrador')
def ordenar_clientes():
    vendedores = Vendedor.query.order_by(Vendedor.nombre).all()
    vendedor = (request.args.get('vendedor') or '').strip()
    ruta = (request.args.get('ruta') or '').strip()

    rutas = []
    clientes = []
    if vendedor:
        rutas = [
            r[0] for r in db.session.query(Cliente.ruta).filter(
                Cliente.codigo_vendedor == vendedor,
                Cliente.ruta.isnot(None), Cliente.ruta != '',
            ).distinct().order_by(Cliente.ruta).all()
        ]
        if ruta:
            clientes = Cliente.query.filter_by(
                codigo_vendedor=vendedor, ruta=ruta
            ).order_by(Cliente.orden_visita.asc(), Cliente.nombre.asc()).all()

    return render_template(
        'clientes/ordenar.html',
        vendedores=vendedores,
        rutas=rutas,
        clientes=clientes,
        filtro_vendedor=vendedor,
        filtro_ruta=ruta,
    )


@clientes_bp.route('/ordenar/guardar', methods=['POST'])
@login_required
@rol_requerido('administrador')
def guardar_orden_clientes():
    data = request.get_json(silent=True) or {}
    try:
        ids = [int(i) for i in (data.get('ids') or [])]
    except (TypeError, ValueError):
        return jsonify({'ok': False, 'error': 'ids inválidos'}), 400
    if not ids:
        return jsonify({'ok': False, 'error': 'Lista vacía'}), 400

    clientes = {c.id: c for c in Cliente.query.filter(Cliente.id.in_(ids)).all()}
    for posicion, cid in enumerate(ids, start=1):
        if cid in clientes:
            clientes[cid].orden_visita = posicion
    db.session.commit()
    return jsonify({'ok': True, 'actualizados': len(clientes)})


@clientes_bp.route('/exportar')
@login_required
@rol_requerido('administrador')
def exportar_clientes():
    vendedor_filtro = request.args.get('vendedor', '')
    query = Cliente.query
    if vendedor_filtro:
        query = query.filter_by(codigo_vendedor=vendedor_filtro)
    clientes = query.order_by(Cliente.codigo_vendedor, Cliente.ruta, Cliente.orden_visita).all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        'codigo_cliente', 'nombre', 'nombre_tienda', 'telefono', 'whatsapp', 'direccion',
        'barrio', 'ciudad', 'latitud', 'longitud', 'ruta', 'orden_visita', 'codigo_vendedor', 'activo',
    ])
    for c in clientes:
        writer.writerow([
            c.codigo_cliente, c.nombre, c.nombre_tienda or '', c.telefono or '', c.whatsapp or '',
            c.direccion or '', c.barrio or '', c.ciudad or '',
            float(c.latitud) if c.latitud else '',
            float(c.longitud) if c.longitud else '',
            c.ruta or '', c.orden_visita or '', c.codigo_vendedor,
            '1' if c.activo else '0',
        ])

    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': 'attachment; filename=clientes_export.csv'},
    )


@clientes_bp.route('/plantilla')
@login_required
@rol_requerido('administrador')
def descargar_plantilla():
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Clientes'
    ws.append([
        'nombre', 'nombre_tienda', 'codigo_vendedor', 'telefono', 'whatsapp',
        'direccion', 'barrio', 'ciudad', 'latitud', 'longitud', 'ruta', 'orden_visita',
    ])
    ws.append([
        'Juan Perez', 'Tienda La Esperanza', 'VEN001', '3001234567', '3001234567',
        'Calle 5 # 10-20', 'El Prado', 'Bucaramanga', '7.11930000', '-73.12270000', 'Ruta 1', '1',
    ])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='plantilla_clientes.xlsx',
    )


def _leer_archivo(archivo):
    """Lee un archivo CSV o XLSX y retorna lista de dicts con claves en minúsculas."""
    nombre = archivo.filename.lower()
    if nombre.endswith('.xlsx'):
        from openpyxl import load_workbook
        wb = load_workbook(archivo, read_only=True, data_only=True)
        ws = wb.active
        filas = list(ws.iter_rows(values_only=True))
        if not filas:
            return []
        headers = [str(h).strip().lower() if h is not None else '' for h in filas[0]]
        result = []
        for fila in filas[1:]:
            if all(v is None for v in fila):
                continue
            result.append({
                headers[i]: (str(fila[i]).strip() if fila[i] is not None else '')
                for i in range(len(headers))
            })
        return result
    else:
        contenido = archivo.read().decode('utf-8-sig')
        reader = csv.DictReader(io.StringIO(contenido))
        return [{k.strip().lower(): (v or '').strip() for k, v in row.items()} for row in reader]


def _procesar_fila(fila, num_fila):
    """
    Procesa una fila del archivo de importación.
    Retorna dict con 'resultado': 'creado' | 'actualizado' | 'error'
    y 'motivo' si es error.
    """
    nombre = (fila.get('nombre') or '').strip()
    codigo_vendedor = (fila.get('codigo_vendedor') or '').strip()
    telefono = (fila.get('telefono') or '').strip()

    if not nombre:
        return {'resultado': 'error', 'fila': num_fila, 'nombre': '(vacío)', 'motivo': 'nombre es requerido'}
    if not codigo_vendedor:
        return {'resultado': 'error', 'fila': num_fila, 'nombre': nombre, 'motivo': 'codigo_vendedor es requerido'}

    vendedor = Vendedor.query.filter_by(codigo_vendedor=codigo_vendedor).first()
    if not vendedor:
        return {'resultado': 'error', 'fila': num_fila, 'nombre': nombre, 'motivo': f'vendedor "{codigo_vendedor}" no existe'}

    # El teléfono es la clave única real del cliente (puede cambiar de vendedor o de nombre).
    if telefono:
        existente = Cliente.query.filter_by(telefono=telefono).first()
    else:
        existente = Cliente.query.filter_by(nombre=nombre, codigo_vendedor=codigo_vendedor).first()

    def _safe_float(val):
        try:
            return float(val) if val not in (None, '') else None
        except (ValueError, TypeError):
            return None

    def _safe_int(val):
        try:
            return int(val) if val not in (None, '') else None
        except (ValueError, TypeError):
            return None

    if existente:
        existente.nombre = nombre
        existente.codigo_vendedor = codigo_vendedor
        if fila.get('nombre_tienda'):
            existente.nombre_tienda = fila['nombre_tienda']
        if fila.get('telefono'):
            existente.telefono = fila['telefono']
        if fila.get('whatsapp'):
            existente.whatsapp = fila['whatsapp']
        if fila.get('direccion'):
            existente.direccion = fila['direccion']
        if fila.get('barrio'):
            existente.barrio = fila['barrio']
        if fila.get('ciudad'):
            existente.ciudad = fila['ciudad']
        lat = _safe_float(fila.get('latitud'))
        if lat is not None:
            existente.latitud = lat
        elif existente.latitud is None:
            existente.latitud = LATITUD_PANADERIA
        lng = _safe_float(fila.get('longitud'))
        if lng is not None:
            existente.longitud = lng
        elif existente.longitud is None:
            existente.longitud = LONGITUD_PANADERIA
        if fila.get('ruta'):
            existente.ruta = fila['ruta']
        orden = _safe_int(fila.get('orden_visita'))
        if orden is not None:
            existente.orden_visita = orden
        return {'resultado': 'actualizado'}
    else:
        lat = _safe_float(fila.get('latitud'))
        lng = _safe_float(fila.get('longitud'))
        nuevo = Cliente(
            codigo_cliente=Cliente.siguiente_codigo(),
            nombre=nombre,
            nombre_tienda=fila.get('nombre_tienda') or None,
            codigo_vendedor=codigo_vendedor,
            telefono=fila.get('telefono') or None,
            whatsapp=fila.get('whatsapp') or None,
            direccion=fila.get('direccion') or None,
            barrio=fila.get('barrio') or None,
            ciudad=fila.get('ciudad') or None,
            ruta=fila.get('ruta') or None,
            latitud=lat if lat is not None else LATITUD_PANADERIA,
            longitud=lng if lng is not None else LONGITUD_PANADERIA,
            orden_visita=_safe_int(fila.get('orden_visita')),
            activo=True,
        )
        db.session.add(nuevo)
        db.session.flush()  # necesario para que siguiente_codigo() vea la fila insertada
        return {'resultado': 'creado'}


@clientes_bp.route('/importar', methods=['GET', 'POST'])
@login_required
@rol_requerido('administrador')
def importar_clientes():
    if request.method == 'GET':
        return render_template('clientes/importar.html')

    archivo = request.files.get('archivo')
    if not archivo or not archivo.filename:
        flash('Selecciona un archivo.', 'danger')
        return redirect(url_for('clientes.importar_clientes'))

    nombre_archivo = archivo.filename.lower()
    if not (nombre_archivo.endswith('.csv') or nombre_archivo.endswith('.xlsx')):
        flash('Solo se aceptan archivos .csv o .xlsx', 'danger')
        return redirect(url_for('clientes.importar_clientes'))

    try:
        filas = _leer_archivo(archivo)
    except Exception as e:
        flash(f'Error al leer el archivo: {e}', 'danger')
        return redirect(url_for('clientes.importar_clientes'))

    creados, actualizados, errores = 0, 0, []

    for i, fila in enumerate(filas, start=2):
        resultado = _procesar_fila(fila, i)
        if resultado['resultado'] == 'creado':
            creados += 1
        elif resultado['resultado'] == 'actualizado':
            actualizados += 1
        else:
            errores.append(resultado)

    db.session.commit()

    fallos_url = None
    if errores:
        from flask import current_app
        import os
        upload_folder = os.path.join(current_app.root_path, 'uploads')
        os.makedirs(upload_folder, exist_ok=True)
        ruta_fallos = os.path.join(upload_folder, 'fallos_importacion_clientes.csv')
        with open(ruta_fallos, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=['fila', 'nombre', 'motivo'],
                                    extrasaction='ignore')
            writer.writeheader()
            writer.writerows(errores)
        fallos_url = '/uploads/fallos_importacion_clientes.csv'

    return render_template(
        'clientes/importar.html',
        resultado={'creados': creados, 'actualizados': actualizados, 'errores': errores},
        fallos_url=fallos_url,
    )
